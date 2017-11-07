#!/usr/bin/env python2

import sys
from openpyxl import load_workbook
from subprocess import check_output

def start():
    try:
        db_file, pcap_file = sys.argv[1:]
    except:
        print "Usage: python tcown-files.py ThreatConnectExport.xlsx /path/to/files"
        sys.exit(1)

    return db_file, path

if __name__ == "__main__":
    db_file, path = start()

    wb = load_workbook(filename=db_file, read_only=True)
    ws = wb['Sheet1']
    print "Scanning {0} for signatures from {1}...".format(pcap_file, db_file)

    signature = ws['B2'].value
    test = check_output(['/usr/bin/tshark', '-r', pcap_file, 'tcp', \
    'contains', signature])

    print test
