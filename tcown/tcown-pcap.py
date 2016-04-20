#!/usr/bin/env python2

import sys
from openpyxl import load_workbook
from subprocess import check_output

def start():
    try:
        db_file, pcap_file = sys.argv[1:]
    except:
        print "Usage: python tcown-pcap.py ThreatConnectExport.xlsx Traffic.pcap"
        sys.exit(1)

    return db_file, pcap_file

if __name__ == "__main__":
    db_file, pcap_file = start()

    wb = load_workbook(filename=db_file, read_only=True)
    ws = wb['Sheet1']
    print "Scanning {0} for signatures from {1}...".format(pcap_file, db_file)

    signature = ws['B2'].value
    test = check_output(['/usr/bin/tshark', '-r', pcap_file, 'tcp', \
    'contains', signature])

    print test

# OLD scapy stuff
#     from scapy.all import *
#     traffic = PcapReader(pcap_file)
#     for p in traffic:
#         # this is obviously bad, fix this
#         if DNS in p:
#             if signature in p[DNS].qd.qname:
#                 print "found!!!", p[DNS].qd.qname
